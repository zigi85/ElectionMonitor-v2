"""
convert_excel.py — Convert סקר מנדטים.xlsx to polls_manual.json

The Excel structure (discovered by inspection):
  - Sections delimited by header rows (row with outlet names in cols 1-9)
  - Each section: [title row with embedded date] → [header row] → [party rows...] → [totals row / empty]
  - Outlet names in headers include embedded poll dates ("חדשות 12 4.12")
  - Columns 1-9 are consistently: כאן11, חדשות12, חדשות13, ערוץ14, ישראל היום, ידיעות/YNET, הארץ, וואלה/מעריב, i24
  - Section title dates can appear in any column (not just col 0)
  - Short dates "12.12" need year inference

Usage:
    uv run --python 3.12 --with openpyxl python scripts/convert_excel.py
"""

import json
import re
import argparse
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Column → outlet mapping (fixed structure, cols 1-9)
# ---------------------------------------------------------------------------
FIXED_COL_OUTLETS = {
    1: "kan11",
    2: "news12",
    3: "news13",
    4: "channel14",
    5: "israel_hayom",
    6: "ynet",
    7: "haaretz",
    8: "walla_maariv",
    9: "i24",
}

# Phrases that identify an outlet header row (partial match, lower-case)
OUTLET_PHRASES = [
    "כאן 11", "כאן11",
    "חדשות 12", "חדשות12",
    "חדשות 13", "חדשות13",
    "ערוץ 14", "עכשיו 14", "חדשות 14",
    "ישראל היום",
    "ידיעות",
    "הארץ",
    "וואלה",
    "מעריב",
    "i24",
]

# ---------------------------------------------------------------------------
# Party name → canonical key mapping (partial/exact)
# ---------------------------------------------------------------------------
PARTY_MAP_EXACT = {
    "הליכוד": "likud",
    "ישר גדי איזנקוט": "yashar",
    "ישר": "yashar",
    "גדי איזנקוט": "yashar",
    "נפתלי בנט": "bennett",
    "נפתלי בנט ויאיר לפיד ביחד": "together",
    "בנט ולפיד ביחד": "together",
    "יחד": "together",
    "יחד (בנט + לפיד)": "together",
    "ישראל ביתנו": "yisrael_beiteinu",
    'ש"ס': "shas",
    "שס": "shas",
    "יש עתיד": "yesh_atid",
    "כחול לבן": "blue_and_white",
    "הדמוקרטים": "democrats",
    "יהדות התורה": "utj",
    'יהדות התורה (אגד"ת)': "utj",
    "עוצמה יהודית": "otzma_yehudit",
    "הציונות הדתית": "religious_zionism",
    'חד"ש-תע"ל': "hadash_taal",
    "חדש תעל": "hadash_taal",
    'רע"מ': "raam",
    "ראם": "raam",
    'בל"ד': "balad",
    "בלד": "balad",
    "המילואימניקים - יועז הנדל": "reservists",
    "המילואימניקים": "reservists",
    "יועז הנדל": "reservists",
    "איחוד הרשימה המשותפת": "joint_list",
    "הרשימה המשותפת": "joint_list",
    "הרשימה המשותפת (מאוחדת)": "joint_list",
}

SKIP_PARTIES = {
    'סה"כ', "סהכ", "סך הכל", "עופר וינטר", "סה׳׳כ",
    "מנדטים", "total", "סך", "שם המפלגה", "מפלגה",
}

# ---------------------------------------------------------------------------
# Static metadata
# ---------------------------------------------------------------------------
PARTY_METADATA = {
    "likud": {"name_he": "הליכוד", "leader": "נתניהו", "color": "#003DA5", "bloc": "coalition", "order": 1},
    "otzma_yehudit": {"name_he": "עוצמה יהודית", "leader": "בן גביר", "color": "#F9A825", "bloc": "coalition", "order": 2},
    "shas": {"name_he": 'ש"ס', "leader": "דרעי", "color": "#00897B", "bloc": "coalition", "order": 3},
    "utj": {"name_he": "יהדות התורה", "leader": "גפני", "color": "#1A237E", "bloc": "coalition", "order": 4},
    "religious_zionism": {"name_he": "הציונות הדתית", "leader": "סמוטריץ'", "color": "#E64A19", "bloc": "coalition", "order": 5},
    "together": {"name_he": "יחד (בנט + לפיד)", "leader": "בנט", "color": "#2E7D32", "bloc": "opposition", "order": 6},
    "bennett": {"name_he": "נפתלי בנט", "leader": "בנט", "color": "#2E7D32", "bloc": "opposition", "order": 6,
                "note": "Pre-merger. After 27.04.2026 use 'together'"},
    "yesh_atid": {"name_he": "יש עתיד", "leader": "לפיד", "color": "#FF8F00", "bloc": "opposition", "order": 7,
                  "note": "Merged into 'together' on 27.04.2026"},
    "yashar": {"name_he": "ישר", "leader": "איזנקוט", "color": "#546E7A", "bloc": "opposition", "order": 8},
    "democrats": {"name_he": "הדמוקרטים", "leader": "", "color": "#C2185B", "bloc": "opposition", "order": 9},
    "yisrael_beiteinu": {"name_he": "ישראל ביתנו", "leader": "ליברמן", "color": "#1565C0", "bloc": "opposition", "order": 10},
    "raam": {"name_he": 'רע"מ', "leader": "עבאס", "color": "#388E3C", "bloc": "unaligned", "order": 11},
    "hadash_taal": {"name_he": 'חד"ש-תע"ל', "leader": "", "color": "#C62828", "bloc": "unaligned", "order": 12},
    "joint_list": {"name_he": "הרשימה המשותפת", "leader": "", "color": "#388E3C", "bloc": "unaligned", "order": 11},
    "blue_and_white": {"name_he": "כחול לבן", "leader": "גנץ", "color": "#039BE5", "bloc": "unaligned", "order": 13,
                       "note": "Consistently below threshold"},
    "reservists": {"name_he": "המילואימניקים", "leader": "הנדל", "color": "#795548", "bloc": "unaligned", "order": 14,
                   "note": "Consistently below threshold"},
    "balad": {"name_he": 'בל"ד', "leader": "", "color": "#7B1FA2", "bloc": "unaligned", "order": 15,
              "note": "Usually below threshold"},
}

OUTLET_METADATA = {
    "kan11": {"name": "כאן 11", "order": 1},
    "news12": {"name": "חדשות 12", "order": 2},
    "news13": {"name": "חדשות 13", "order": 3},
    "channel14": {"name": "ערוץ 14", "order": 4},
    "israel_hayom": {"name": "ישראל היום", "order": 5},
    "ynet": {"name": "ידיעות אחרונות", "order": 6},
    "haaretz": {"name": "הארץ", "order": 7},
    "walla": {"name": "וואלה", "order": 8},
    "maariv": {"name": "מעריב", "order": 9},
    "walla_maariv": {"name": "וואלה/מעריב", "order": 8, "note": "Combined column"},
    "i24": {"name": "i24", "order": 10},
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def norm(v) -> str:
    """Normalize a cell value to a clean string."""
    if v is None:
        return ""
    return str(v).strip().replace("‏", "").replace("‎", "").replace("\xa0", " ").replace("\n", " ").strip()


def row_text(row) -> list[str]:
    return [norm(c) for c in row]


def is_header_row(row) -> bool:
    """True if ≥2 cells in cols 1+ contain known outlet phrases."""
    texts = row_text(row)
    hits = 0
    for i, t in enumerate(texts):
        if i == 0:
            continue
        tl = t.lower()
        for phrase in OUTLET_PHRASES:
            if phrase.lower() in tl:
                hits += 1
                break
    return hits >= 2


def extract_date_from_text(text: str) -> tuple[str, str] | None:
    """
    Find a date within a string.
    Handles: "30.05.2026", "12.12", range "03-04.2026", month.year "04.2026".
    Avoids matching percentages like "0 (1.4%)" or "2.6%".
    Returns (iso, label) or None.
    """
    text = text.strip()
    if not text:
        return None

    # 1. Full date: D.M.YYYY
    m = re.search(r'\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b', text)
    if m:
        try:
            d = datetime.strptime(f"{m.group(1)}.{m.group(2)}.{m.group(3)}", "%d.%m.%Y")
            return (d.date().isoformat(), d.strftime("%d.%m.%Y"))
        except ValueError:
            pass

    # 2. Month range or plain month.year: "03-04.2026" or "04.2026"
    #    Extracts the last 2-digit month before ".YYYY"
    m = re.search(r'\b(?:\d{2}-)?(\d{2})\.(\d{4})\b', text)
    if m:
        try:
            month, year = int(m.group(1)), int(m.group(2))
            if 1 <= month <= 12 and 2020 <= year <= 2030:
                d = date(year, month, 1)
                return (d.isoformat(), d.strftime("01.%m.%Y"))
        except ValueError:
            pass

    # 3. Short date: D.M — but NOT inside "(...%)" percentage context
    #    Negative lookbehind for "(" guards "0 (1.4%)" → skips "1.4"
    #    Negative lookahead for "%" guards "2.6%"
    m = re.search(r'(?<!\()\b(\d{1,2})\.(\d{1,2})\b(?!%)', text)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        if 1 <= day <= 31 and 1 <= month <= 12:
            year = 2025 if month >= 11 else 2026
            try:
                d = date(year, month, day)
                return (d.isoformat(), d.strftime("%d.%m.%Y"))
            except ValueError:
                pass

    return None


def find_section_date(rows: list, header_idx: int) -> tuple[str, str]:
    """
    Find the date for a section. Strategy:
    1. Scan backwards up to 12 rows for a date in any cell (title rows may be in any column).
    2. Fallback: extract date from outlet column headers like "חדשות 13 (20.1)".
    Returns (iso, label).
    """
    # Pass 1: scan backwards for a title row date (any cell, any column)
    for i in range(header_idx - 1, max(-1, header_idx - 12), -1):
        row = rows[i]
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, datetime):
                d = cell.date()
                return (d.isoformat(), d.strftime("%d.%m.%Y"))
            result = extract_date_from_text(str(cell))
            if result:
                return result

    # Pass 2: extract date from outlet header cells, e.g. "חדשות 13 (20.1)"
    header_row = rows[header_idx]
    for cell in header_row[1:]:  # skip col 0 (party name column)
        if cell is None:
            continue
        m = re.search(r'\((\d{1,2})\.(\d{1,2})(?:\.(\d{4}))?\)', str(cell))
        if m:
            day, month = int(m.group(1)), int(m.group(2))
            year = int(m.group(3)) if m.group(3) else (2025 if month >= 11 else 2026)
            if 1 <= day <= 31 and 1 <= month <= 12:
                try:
                    d = date(year, month, day)
                    return (d.isoformat(), d.strftime("%d.%m.%Y"))
                except ValueError:
                    pass

    return ("unknown", "לא ידוע")


def parse_party(name: str):
    """Map a Hebrew party name string to a canonical key, or None to skip."""
    n = norm(name)
    if not n:
        return None
    # Skip known non-party rows
    if any(s in n for s in SKIP_PARTIES):
        return None
    # Exact match
    key = PARTY_MAP_EXACT.get(n)
    if key:
        return key
    # Substring match (handles slight variations)
    for hebrew, k in PARTY_MAP_EXACT.items():
        if hebrew in n or n in hebrew:
            return k
    return None


def parse_seat(raw) -> tuple[int, float | None]:
    """Returns (seats, below_threshold_pct or None)."""
    if raw is None or raw == "":
        return (0, None)
    if isinstance(raw, (int, float)):
        v = int(raw)
        return (v, None)
    s = norm(str(raw))
    if not s:
        return (0, None)
    # "0 (2.6%)"
    m = re.match(r"^(\d+)\s*\((\d+\.?\d*)%\)$", s)
    if m:
        return (int(m.group(1)), float(m.group(2)))
    try:
        return (int(float(s)), None)
    except ValueError:
        return (0, None)


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_excel(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Sheet: {ws.title!r}  rows={len(rows)}  cols={ws.max_column}")

    # --- Find all header row indices ---
    header_indices = [i for i, row in enumerate(rows) if is_header_row(row)]
    print(f"  Found {len(header_indices)} section header rows at indices: {header_indices[:20]}...")

    if not header_indices:
        raise ValueError("No header rows detected. Inspect the file with inspect_excel.py.")

    # --- Parse each section ---
    # Sections run from header_row+1 to the next header_row (exclusive)
    raw_sections: list[dict] = []

    for sec_num, hdr_idx in enumerate(header_indices):
        end_idx = header_indices[sec_num + 1] if sec_num + 1 < len(header_indices) else len(rows)

        # Find date by scanning backwards from header row
        iso_date, label = find_section_date(rows, hdr_idx)

        # Parse party rows in this section
        outlet_data: dict[str, dict] = {oid: {} for oid in FIXED_COL_OUTLETS.values()}
        below_threshold: dict[str, dict] = {oid: {} for oid in FIXED_COL_OUTLETS.values()}

        for row in rows[hdr_idx + 1 : end_idx]:
            texts = row_text(row)
            party_name = texts[0] if texts else ""

            party_key = parse_party(party_name)
            if party_key is None:
                continue

            for col_idx, outlet_id in FIXED_COL_OUTLETS.items():
                if col_idx >= len(row):
                    continue
                seats, pct = parse_seat(row[col_idx])
                if seats > 0:
                    outlet_data[outlet_id][party_key] = seats
                if pct is not None:
                    below_threshold[outlet_id][party_key] = pct

        raw_sections.append({
            "date": iso_date,
            "label": label,
            "outlet_data": outlet_data,
            "below_threshold": below_threshold,
        })
        print(f"  Section {sec_num+1:2d}: {label}  outlets_with_data={sum(1 for d in outlet_data.values() if d)}")

    # --- Merge sections with the same date ---
    merged: dict[str, dict] = {}  # date → merged section

    for sec in raw_sections:
        d = sec["date"]
        if d not in merged:
            merged[d] = {"date": d, "label": sec["label"], "outlet_data": {}, "below_threshold": {}}
        for oid, parties in sec["outlet_data"].items():
            if parties:
                # Later sections for same date override earlier ones per outlet
                merged[d]["outlet_data"][oid] = parties
        for oid, pcts in sec["below_threshold"].items():
            if pcts:
                merged[d]["below_threshold"][oid] = pcts

    # --- Build final timestamps array ---
    timestamps = []
    for d in sorted(merged):
        ms = merged[d]
        polls = []
        for oid in ["kan11", "news12", "news13", "channel14", "israel_hayom", "ynet", "haaretz", "walla_maariv", "i24"]:
            parties = ms["outlet_data"].get(oid, {})
            if not parties:
                continue
            total = sum(parties.values())
            poll: dict = {
                "outlet": OUTLET_METADATA.get(oid, {}).get("name", oid),
                "outlet_id": oid,
                "parties": parties,
                "total": total,
            }
            btp = ms["below_threshold"].get(oid, {})
            if btp:
                poll["below_threshold_pcts"] = btp

            if oid == "walla_maariv":
                # Split into two identical entries
                for split_id, split_name in [("walla", "וואלה"), ("maariv", "מעריב")]:
                    p = dict(poll)
                    p["outlet"] = split_name
                    p["outlet_id"] = split_id
                    polls.append(p)
            else:
                polls.append(poll)

        if polls:
            timestamps.append({
                "id": ms["date"],
                "label": ms["label"],
                "polls": polls,
            })

    print(f"\n  Total timestamps (after merge): {len(timestamps)}")

    last_updated = timestamps[-1]["id"] if timestamps else date.today().isoformat()
    return {
        "last_updated": last_updated,
        "timestamps": timestamps,
        "party_metadata": PARTY_METADATA,
        "outlet_metadata": OUTLET_METADATA,
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="public/data/סקר מנדטים.xlsx")
    ap.add_argument("--output", default="public/data/polls_manual.json")
    args = ap.parse_args()

    inp = Path(args.input)
    out = Path(args.output)

    if not inp.exists():
        print(f"ERROR: {inp} not found")
        sys.exit(1)

    print(f"Parsing: {inp}")
    data = parse_excel(inp)

    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {len(data['timestamps'])} timestamps to: {out}")
    print("\n--- Timestamp summary ---")
    for ts in data["timestamps"]:
        outlets = [p["outlet_id"] for p in ts["polls"]]
        total_outlets = len(outlets)
        print(f"  {ts['label']}: {total_outlets} outlet(s) — {', '.join(outlets)}")


if __name__ == "__main__":
    main()
